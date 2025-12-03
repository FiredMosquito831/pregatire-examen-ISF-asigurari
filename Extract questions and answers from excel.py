import openpyxl
import json

# 1. Load the workbook
file_path = 'data.xlsx'  # Make sure this matches your file name
wb = openpyxl.load_workbook(file_path)
ws = wb.active

questions_list = []

# 2. Iterate through the sheet in steps of 3
# We start at row 2 (assuming headers) to the end of the sheet
for i in range(2, ws.max_row + 1, 3):

    # Extract ID and Question from the top-left of the merged block
    q_id = ws.cell(row=i, column=1).value
    q_text = ws.cell(row=i, column=2).value

    # Stop if we hit an empty row
    if q_id is None:
        break

    answers_list = []

    # 3. Iterate through the 3 individual answer rows for this set
    for offset in range(3):
        current_row = i + offset
        cell_val = ws.cell(row=current_row, column=3).value

        if cell_val:
            # Convert to string to avoid errors with numbers
            text_val = str(cell_val)

            # Check if '*' exists in the text
            if "*" in text_val:
                is_correct = True
                # Remove the asterisk and whitespace for clean output
                clean_text = text_val.replace("*", "").strip()
            else:
                is_correct = False
                clean_text = text_val.strip()

            # Create the answer object
            ans_obj = {
                "text": clean_text,
                "isCorrect": is_correct
            }
            answers_list.append(ans_obj)

    # 4. Build the Question Set Object
    q_obj = {
        "id": q_id,
        "question": q_text,
        "answers": answers_list
    }

    questions_list.append(q_obj)

# 5. Dump to actual JSON file
output_filename = 'output.json'
with open(output_filename, 'w', encoding='utf-8') as f:
    json.dump(questions_list, f, indent=4, ensure_ascii=False)

print(f"Successfully processed {len(questions_list)} question sets into {output_filename}.")